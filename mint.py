#!/usr/bin/env python

#imports
import openpyxl
import os


# Workhorse Functions

# Purpose: Take new data from the mint workbook, and transfer it to the destination workbook
# Notes: Prepend the newer entries, ignore repeat dates
def mintPull(mintBook, destBook):

	# Get basic solution working
	mintSheet = mintBook.active
	destSheet = destBook.active

	# Copy the title row (row 1) to the top of the doc
	for titleRow in mintSheet.iter_rows(max_row=1):
		for cellIndex in range(mintSheet.max_column):
			destSheet.cell(row=1,column= cellIndex + 1) = titleRow[cellIndex].value

	# Copy the rest of the rows if the date being parsed is equal to or greater
	for cellRow in mintSheet.iter_rows(min_row=2):
		print(cellRow[0].value)
		for cellIndex in range(mintSheet.max_column):
			if isDateGreater(cellRow[0].value, destSheet.cell(row=,column=).value):
			#	destSheet.insert_rows(1, amount=1) # Pre-pend blank row
			#	destSheet.cell(row=2, column= cellIndex + 2).value = cellRow[cellIndex].value
			#	print( destSheet.cell(row=2, column= cellIndex + 2).value )
				pass

# Parse through the dates and append numbers if duplicates present
def labelDates(book):



# Helper Functions
def initWorkbook(src):
	workbookObj = openpyxl.load_workbook(src)
	return workbookObj

def isDateGreater(date, date2):
	if dateParse(date) >= dateParse(date2) or dateParse(date2) == None:
		return True
	else:
		return False

def dateParse(date):
	string = ""
	for letter in str(date):
		if letter.isdigit():
			string += letter
	return int(string)


# Create a set based on the Excel Column Name passed in
def listFromColumn(sheet, column):
	colList = []
	for columnCell in sheet[column]:
		if not columnCell.value in colList:
			colList.append(str(columnCell.value))
	return colList

# Print the rows for the worksheet
def print_rows(sheet, isPrint=True):
	print("==============\nSheet Report for: ", sheet.title)
	print("Rows:", sheet.max_row, "Cols:", sheet.max_column)

	i = 0
	if isPrint == True:
		for row in sheet.iter_rows(values_only=True):
			print(f'Line {i}:', row[0])
			i += 1

# Driver Code
if __name__ == '__main__':
	# Universal Test Flags
	printFlag = True
	saveFlag = True

	# get CLI arguments <python3> <programName> <mintsrc> <destsrc>
	#mintsrc = sys.argv[1]
	#destsrc = sys.argv[2]
	mintsrc = 'test/mintTest.xlsx'
	destsrc = 'test/output.xlsx'

	mintBook = initWorkbook(mintsrc)
	destBook = initWorkbook(destsrc)

	# Parse the dates, append numbers to better distinguish between duplicate dates


	# Pull in Mint data into the new workbook
	mintPull(mintBook, destBook)

	#print_rows(destBook.active)
	if saveFlag:
		destBook.save('test/output.xlsx')
		pass
	print("Done")
